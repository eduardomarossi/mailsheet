from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import re


class InvalidSheetRangeException(Exception):
    pass


def get_ranges(sheet_range):
    min_col, max_col, min_row, max_row = None, None, None, None
    p = re.compile('([A-Za-z]+)([0-9]+):([A-za-z]+)([0-9]+)')
    m = p.match(sheet_range)
    if m:
        min_col = column_index_from_string(m.group(1))
        min_row = int(m.group(2))
        max_col = column_index_from_string(m.group(3))
        max_row = int(m.group(4))
        return min_col, max_col, min_row, max_row

    p = re.compile('([A-Za-z]+):([A-za-z]+)')
    m = p.match(sheet_range)
    if m:
        min_col = column_index_from_string(m.group(1))
        max_col = column_index_from_string(m.group(2))
        return min_col, max_col, min_row, max_row

    raise InvalidSheetRangeException('Provided range is invalid')


def read_sheet(file_path, sheet_name, sheet_range):
    wb = load_workbook(filename=file_path)
    ws = wb[sheet_name]
    min_col, max_col, min_row, max_row = get_ranges(sheet_range)
    cells = ws.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)

    values = []
    for r in cells:
        values_row = []
        for c in r:
            val = c.value
            if val is None:
                val = ''

            if type(val) is datetime:
                val = val.strftime('%d/%m/%Y')
            values_row.append(val)

        values.append(values_row)

    return values


def open_sheet_keep_row(file_path, dest_path, sheet_name, starts_at, row_index):
    wb = load_workbook(filename=file_path)
    sheets = wb.sheetnames

    ## delete not using sheets
    for s in sheets:
        if s != sheet_name:
            wb.remove(wb.get_sheet_by_name(s))

    ws = wb[sheet_name]

    # wanna keep this
    real_index = starts_at + row_index

    # header: 1-3 data: 4  row:6
    # (4, 2)
    if real_index > starts_at:
        ws.delete_rows(starts_at, row_index)

    ws.delete_rows(starts_at+1, ws.max_row)

    wb.save(dest_path)
    wb.close()




if __name__ == '__main__':
    open_sheet_keep_row('teste2.xlsx', 'teste3.xlsx', 'Planilha1', 2, 1)


